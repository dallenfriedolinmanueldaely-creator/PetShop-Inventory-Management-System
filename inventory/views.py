import csv
import datetime
from datetime import date
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db import transaction, IntegrityError
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from .decorators import login_required_custom, admin_required, staff_required, get_user_role
from .models import Product, StockIn, StockOut, TransactionHistory, ActivityLog, UserProfile
from .forms import ProductForm, UserCreateForm, UserEditForm, StockInForm, StockOutForm
from .constants import LOW_STOCK_THRESHOLD
from .algorithms import linear_search, binary_search_by_nama, bubble_sort, selection_sort, insertion_sort

def _log_activity(user, activity_type, product=None, details=''):
    role = get_user_role(user) if user else ''
    if not role:
        role = ''
    username = user.username if user else ''
    product_name = product.name if product else ''
    ActivityLog.objects.create(
        user=user,
        username=username,
        user_role=role,
        activity_type=activity_type,
        product=product,
        product_name=product_name,
        details=details
    )

def _apply_product_filters(request, products):
    from .algorithms import get_algorithm_info

    keyword = request.GET.get('q', '').strip()
    search_field = request.GET.get('search_field', 'name')
    filter_pet = request.GET.get('pet', '')
    filter_category = request.GET.get('category', '')
    sort_by = request.GET.get('sort_by', '')
    sort_order = request.GET.get('sort_order', 'asc')
    search_algo = request.GET.get('search_algo', 'linear')
    sort_algo = request.GET.get('sort_algo', 'bubble')

    if filter_pet:
        products = products.filter(pet=filter_pet)
    if filter_category:
        products = products.filter(category=filter_category)

    if keyword:
        if search_algo == 'binary' and search_field == 'name':
            products = binary_search_by_nama(products, keyword)
        else:
            products = linear_search(products, keyword, field=search_field)

    if sort_by:
        ascending = (sort_order == 'asc')
        if sort_algo == 'selection':
            products = selection_sort(products, field=sort_by, ascending=ascending)
        elif sort_algo == 'insertion':
            products = insertion_sort(products, field=sort_by, ascending=ascending)
        else:
            products = bubble_sort(products, field=sort_by, ascending=ascending)

    search_result_count = len(products) if isinstance(products, list) else products.count()

    algo_info_search = None
    if keyword:
        algo_search_key = 'linear_search' if search_algo == 'linear' or search_field != 'name' else 'binary_search'
        algo_info_search = get_algorithm_info(algo_search_key)

    algo_info_sort = None
    if sort_by:
        algo_info_sort = get_algorithm_info(f'{sort_algo}_sort')

    filter_ctx = {
        'keyword': keyword,
        'search_field': search_field,
        'filter_pet': filter_pet,
        'filter_category': filter_category,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'search_algo': search_algo,
        'sort_algo': sort_algo,
        'search_result_count': search_result_count,
        'algo_info_search': algo_info_search,
        'algo_info_sort': algo_info_sort,
    }
    return products, filter_ctx


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            _log_activity(user, 'LOGIN', details=f'User {user.username} logged in.')
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password!')

    return render(request, 'inventory/login.html', {'page_title': 'Login'})

@require_POST
def logout_view(request):
    if request.user.is_authenticated:
        _log_activity(request.user, 'LOGOUT', details=f'User {request.user.username} logged out.')
    logout(request)
    return redirect('login')

@login_required_custom
def dashboard(request):
    stats = Product.objects.aggregate(
        total_products=Count('id'),
        total_stock=Sum('stock'),
        out_of_stock=Count('id', filter=Q(stock=0)),
        low_stock=Count('id', filter=Q(stock__gt=0, stock__lte=LOW_STOCK_THRESHOLD)),
        total_cats=Count('id', filter=Q(pet='Cat')),
        total_dogs=Count('id', filter=Q(pet='Dog'))
    )
    total_products = stats['total_products'] or 0
    total_stock = stats['total_stock'] or 0
    out_of_stock = stats['out_of_stock'] or 0
    low_stock = stats['low_stock'] or 0
    total_cats = stats['total_cats'] or 0
    total_dogs = stats['total_dogs'] or 0

    recent_transactions = TransactionHistory.objects.select_related('product', 'recorded_by').order_by('-date')[:10]
    latest_products = Product.objects.order_by('-created_at')[:6]

    category_data = Product.objects.values('category').annotate(count=Count('category')).order_by('category')

    today = timezone.now()
    today_date = today.date()
    start_of_month = timezone.make_aware(datetime.datetime(today.year, today.month, 1))

    stock_in_today = StockIn.objects.filter(date__date=today_date).count()
    stock_out_today = StockOut.objects.filter(date__date=today_date).count()

    stock_in_this_month = StockIn.objects.filter(date__gte=start_of_month).aggregate(Sum('quantity'))['quantity__sum'] or 0
    stock_out_this_month = StockOut.objects.filter(date__gte=start_of_month).aggregate(Sum('quantity'))['quantity__sum'] or 0

    seven_days_ago = today_date - datetime.timedelta(days=6)

    in_data = StockIn.objects.filter(date__date__gte=seven_days_ago) \
                             .values('date__date') \
                             .annotate(total=Sum('quantity'))
    in_map = {item['date__date']: item['total'] for item in in_data if item['date__date']}

    out_data = StockOut.objects.filter(date__date__gte=seven_days_ago) \
                              .values('date__date') \
                              .annotate(total=Sum('quantity'))
    out_map = {item['date__date']: item['total'] for item in out_data if item['date__date']}

    movement_labels = []
    movement_in = []
    movement_out = []
    for i in range(6, -1, -1):
        d = today_date - datetime.timedelta(days=i)
        movement_labels.append(d.strftime('%b %d'))
        movement_in.append(int(in_map.get(d, 0)))
        movement_out.append(int(out_map.get(d, 0)))

    cat_labels = [item['category'] for item in category_data]
    cat_counts = [item['count'] for item in category_data]

    total_users = User.objects.count()
    team_members = User.objects.select_related('profile').all()[:6]
    recent_activity = ActivityLog.objects.select_related('user__profile', 'product').order_by('-timestamp')[:8]

    low_stock_products = Product.objects.filter(
        stock__gt=0, stock__lte=LOW_STOCK_THRESHOLD
    ).order_by('stock')[:5]
    out_of_stock_products = Product.objects.filter(stock=0).order_by('name')[:5]

    context = {
        'total_products': total_products,
        'total_stock': total_stock,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'total_cats': total_cats,
        'total_dogs': total_dogs,
        'recent_transactions': recent_transactions,
        'latest_products': latest_products,
        'category_data': list(category_data),
        'stock_in_today': stock_in_today,
        'stock_out_today': stock_out_today,
        'stock_in_this_month': stock_in_this_month,
        'stock_out_this_month': stock_out_this_month,
        'movement_labels': movement_labels,
        'movement_in': movement_in,
        'movement_out': movement_out,
        'cat_labels': cat_labels,
        'cat_counts': cat_counts,
        'total_users': total_users,
        'team_members': team_members,
        'recent_activity': recent_activity,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'user_role': get_user_role(request.user),
        'page_title': 'Dashboard',
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required_custom
def product_list(request):
    products = Product.objects.all()
    products, filter_ctx = _apply_product_filters(request, products)

    low_stock_count = Product.objects.filter(stock__gt=0, stock__lte=LOW_STOCK_THRESHOLD).count()
    out_of_stock_count = Product.objects.filter(stock=0).count()

    categories = [c[0] for c in Product.CATEGORY_CHOICES]

    paginator = Paginator(products, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'categories': categories,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'user_role': get_user_role(request.user),
        'page_title': 'Product List',
        **filter_ctx,
    }
    return render(request, 'inventory/product_list.html', context)

@login_required_custom
def low_stock_alert(request):
    low_stock_products = Product.objects.filter(
        stock__gt=0, stock__lte=LOW_STOCK_THRESHOLD
    ).order_by('stock')
    out_of_stock_products = Product.objects.filter(stock=0).order_by('name')

    context = {
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'threshold': LOW_STOCK_THRESHOLD,
        'user_role': get_user_role(request.user),
        'page_title': 'Low Stock Alert',
    }
    return render(request, 'inventory/low_stock_alert.html', context)

@admin_required
def export_products_csv(request):
    products = Product.objects.all()
    products, filter_ctx = _apply_product_filters(request, products)

    response = HttpResponse(content_type='text/csv')
    filename = f'products_export_{date.today().isoformat()}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Pet', 'Category', 'Price (Rp)', 'Stock', 'Status', 'Date Added'])

    for p in products:
        writer.writerow([
            p.pk, p.name, p.pet, p.category,
            p.price, p.stock, p.stock_status,
            p.date_added.strftime('%Y-%m-%d') if p.date_added else '',
        ])

    _log_activity(request.user, 'EXPORT_CSV',
                  details=f'Exported products to CSV by {request.user.username}')
    return response

@admin_required
def export_products_xlsx(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export failed: the "openpyxl" library is not installed.')
        return redirect('product_list')

    products = Product.objects.all()
    products, filter_ctx = _apply_product_filters(request, products)
    products_list = list(products)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    header_fill = PatternFill('solid', fgColor='8B5E3C')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center')

    alt_fill = PatternFill('solid', fgColor='FFF8F0')
    thin = Side(style='thin', color='D0C0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    low_fill = PatternFill('solid', fgColor='FFF3CD')
    out_fill = PatternFill('solid', fgColor='F8D7DA')

    headers = ['ID', 'Name', 'Pet', 'Category', 'Price (Rp)', 'Stock', 'Status', 'Date Added']
    col_widths = [6, 36, 8, 14, 16, 8, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, p in enumerate(products_list, start=2):
        row_data = [
            p.pk, p.name, p.pet, p.category,
            float(p.price), p.stock, p.stock_status,
            p.date_added.strftime('%Y-%m-%d') if p.date_added else '',
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        if p.stock == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = out_fill
        elif p.stock <= LOW_STOCK_THRESHOLD:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = low_fill

    ws_sum = wb.create_sheet('Summary')
    summary_data = [
        ('Export Date', date.today().isoformat()),
        ('Total Products', len(products_list)),
        ('Low Stock (≤5)', sum(1 for p in products_list if 0 < p.stock <= LOW_STOCK_THRESHOLD)),
        ('Out of Stock', sum(1 for p in products_list if p.stock == 0)),
    ]
    for r, (label, value) in enumerate(summary_data, start=1):
        ws_sum.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_sum.cell(row=r, column=2, value=value)
    ws_sum.column_dimensions['A'].width = 22
    ws_sum.column_dimensions['B'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'products_export_{date.today().isoformat()}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    _log_activity(request.user, 'EXPORT_XLSX',
                  details=f'Exported products to Excel by {request.user.username}')
    return response

@admin_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            _log_activity(request.user, 'PRODUCT_ADDED', product=product,
                          details=f'Added product: {product.name}')
            messages.success(request, 'Product added successfully.')
            return redirect('product_list')
    else:
        form = ProductForm()

    context = {
        'form': form,
        'user_role': get_user_role(request.user),
        'page_title': 'Add Product',
        'action': 'Add',
    }
    return render(request, 'inventory/product_form.html', context)

@admin_required
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            _log_activity(request.user, 'PRODUCT_UPDATED', product=product,
                          details=f'Updated product: {product.name}')
            messages.success(request, 'Product updated successfully.')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)

    context = {
        'form': form,
        'product': product,
        'user_role': get_user_role(request.user),
        'page_title': f'Edit: {product.name}',
        'action': 'Edit',
    }
    return render(request, 'inventory/product_form.html', context)

@admin_required
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        name = product.name
        try:
            with transaction.atomic():
                _log_activity(request.user, 'PRODUCT_DELETED', product=product,
                              details=f'Deleted product: {name}')
                product.delete()
            messages.success(request, 'Product deleted successfully.')
            return redirect('product_list')
        except ProtectedError:
            messages.error(
                request,
                f"Cannot delete '{name}' because it has transaction history records associated with it."
            )
            return redirect('product_list')

    context = {
        'product': product,
        'user_role': get_user_role(request.user),
        'page_title': f'Delete: {product.name}',
    }
    return render(request, 'inventory/delete_product.html', context)

@login_required_custom
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    history = TransactionHistory.objects.filter(product=product).select_related('recorded_by').order_by('-date')[:20]

    context = {
        'product': product,
        'history': history,
        'user_role': get_user_role(request.user),
        'page_title': f'Detail: {product.name}',
    }
    return render(request, 'inventory/product_detail.html', context)

@staff_required
def stock_in_list(request):
    records = StockIn.objects.select_related('product', 'recorded_by').order_by('-date')

    filter_pet = request.GET.get('pet', '')
    if filter_pet in ('Cat', 'Dog'):
        records = records.filter(product__pet=filter_pet)

    paginator = Paginator(records, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'filter_pet': filter_pet,
        'user_role': get_user_role(request.user),
        'page_title': 'Stock In',
    }
    return render(request, 'inventory/stock_in_list.html', context)

@staff_required
def stock_in_add(request):
    if request.method == 'POST':
        form = StockInForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            record.recorded_by = request.user
            record.save()
            messages.success(request, 'Stock added successfully.')
            return redirect('stock_in_list')
    else:
        form = StockInForm()

    context = {
        'form': form,
        'user_role': get_user_role(request.user),
        'page_title': 'Add Stock In',
    }
    return render(request, 'inventory/stock_in_form.html', context)

@staff_required
def stock_out_list(request):
    records = StockOut.objects.select_related('product', 'recorded_by').order_by('-date')

    filter_pet = request.GET.get('pet', '')
    if filter_pet in ('Cat', 'Dog'):
        records = records.filter(product__pet=filter_pet)

    paginator = Paginator(records, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'filter_pet': filter_pet,
        'user_role': get_user_role(request.user),
        'page_title': 'Stock Out',
    }
    return render(request, 'inventory/stock_out_list.html', context)

@staff_required
def stock_out_add(request):
    if request.method == 'POST':
        form = StockOutForm(request.POST)
        if form.is_valid():
            try:
                record = form.save(commit=False)
                record.recorded_by = request.user
                record.save()
                messages.success(request, 'Stock deducted successfully.')
                return redirect('stock_out_list')
            except ValueError as e:
                messages.error(request, f"Transaction Error: {str(e)}")
            except IntegrityError:
                messages.error(request, "Transaction Error: Insufficient stock or concurrent transaction error.")
        else:
            messages.error(request, "Failed to record Stock Out. Please correct the errors in the form.")
    else:
        form = StockOutForm()

    context = {
        'form': form,
        'user_role': get_user_role(request.user),
        'page_title': 'Add Stock Out',
    }
    return render(request, 'inventory/stock_out_form.html', context)

@admin_required
def activity_log(request):
    logs = ActivityLog.objects.select_related('user__profile', 'product').order_by('-timestamp')

    filter_type = request.GET.get('type', '')
    if filter_type:
        logs = logs.filter(activity_type=filter_type)

    filter_user = request.GET.get('user', '')
    if filter_user:
        logs = logs.filter(Q(username__icontains=filter_user) | Q(user__username__icontains=filter_user))

    paginator = Paginator(logs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    activity_types = ActivityLog.ACTIVITY_CHOICES

    context = {
        'page_obj': page_obj,
        'filter_type': filter_type,
        'filter_user': filter_user,
        'activity_types': activity_types,
        'user_role': get_user_role(request.user),
        'page_title': 'Activity Log',
    }
    return render(request, 'inventory/activity_log.html', context)

@admin_required
def user_list(request):
    users = User.objects.select_related('profile').order_by('username')
    context = {
        'users': users,
        'user_role': get_user_role(request.user),
        'page_title': 'User Management',
    }
    return render(request, 'inventory/user_list.html', context)

@admin_required
def user_add(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            user = form.save()
            role = form.cleaned_data.get('role', 'staff')
            photo = form.cleaned_data.get('photo')
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = role
            if photo:
                profile.photo = photo
            profile.save()

            from django.contrib.auth.models import Group
            user.groups.clear()
            group_name = role.capitalize()
            group, _ = Group.objects.get_or_create(name=group_name)
            user.groups.add(group)

            messages.success(request, 'User added successfully.')
            return redirect('user_list')
    else:
        form = UserCreateForm(user=request.user)

    context = {
        'form': form,
        'user_role': get_user_role(request.user),
        'page_title': 'Add User',
    }
    return render(request, 'inventory/user_form.html', context)

@admin_required
def user_edit(request, pk):
    target_user = get_object_or_404(User, pk=pk)
    profile, _ = UserProfile.objects.get_or_create(user=target_user)

    requester_role = get_user_role(request.user)
    target_role = get_user_role(target_user)
    if requester_role == 'admin' and target_role == 'owner':
        messages.error(request, "Admins are not allowed to edit the Owner's account.")
        return redirect('user_list')

    if request.method == 'POST':
        form = UserEditForm(request.POST, request.FILES, instance=target_user,
                            initial={'role': profile.role}, user=request.user)
        if form.is_valid():
            user = form.save()
            ActivityLog.objects.filter(user=user).update(username=user.username)
            role = form.cleaned_data.get('role', 'staff')
            photo = form.cleaned_data.get('photo')
            profile.role = role
            if photo:
                profile.photo = photo
            profile.save()

            from django.contrib.auth.models import Group
            user.groups.clear()
            group_name = role.capitalize()
            group, _ = Group.objects.get_or_create(name=group_name)
            user.groups.add(group)

            messages.success(request, 'User updated successfully.')
            return redirect('user_list')
    else:
        form = UserEditForm(instance=target_user, initial={'role': profile.role}, user=request.user)

    context = {
        'form': form,
        'target_user': target_user,
        'profile': profile,
        'user_role': get_user_role(request.user),
        'page_title': f'Edit User: {target_user.username}',
    }
    return render(request, 'inventory/user_form.html', context)

@admin_required
def user_delete(request, pk):
    target_user = get_object_or_404(User, pk=pk)

    if target_user == request.user:
        return redirect('user_list')

    requester_role = get_user_role(request.user)
    target_role = get_user_role(target_user)
    if requester_role == 'admin' and target_role == 'owner':
        return redirect('user_list')

    if request.method == 'POST':
        target_user.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('user_list')

    context = {
        'target_user': target_user,
        'user_role': get_user_role(request.user),
        'page_title': f'Delete User: {target_user.username}',
    }
    return render(request, 'inventory/delete_user.html', context)
